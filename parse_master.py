"""마스터 엑셀(2026 AI.xlsx) 매핑 로더.

3개 블록(본부별 공용 / 부설연구소 / 전직원)을 인식해
이메일(또는 ID) → 사용자표기 를 만든다.
사용자표기 규칙:
  - 본부별 공용 계정 → 본부명
  - 부설연구소 계정 → "부설연구소 " + 사람이름
  - 전직원 → "회사공용"
시트에 이메일이 없는 부설연구소 개인은 id_aliases(별칭)로 보완.
노랑 형광펜(FFFFFF00) = Claude Max = 기안서 대상.
"""
import sys
import json
import re
import openpyxl

YELLOW = "FFFFFF00"

# 시트에 이메일이 없는 개인/별칭 보완 (사용자 제공 단서)
DEFAULT_ALIASES = {
    "towyssl": "부설연구소 김은주 대리",
    "mid0701": "부설연구소 김은주 대리",
    "jongwoo": "부설연구소 박종우 소장",
    "aumlee0101": "설계1사업부 1본부",
    "aumlee0401": "설계2사업부 1본부",
    "aumlee0301": "해외사업본부",
    "aumlee0501": "설계2사업부 주거사업본부",
}


def _is_yellow(cell):
    f = cell.fill
    return bool(f and f.patternType and getattr(f.fgColor, "rgb", None) == YELLOW)


def load_master(xlsx, aliases=None):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    section = None          # 1=본부공용, 2=부설연구소, 3=전직원
    cur_hq = ""             # 본부명(블록1 forward-fill)
    cur_name = ""           # 사람이름(블록2 forward-fill)
    email_to_user = {}
    name_to_user = {}
    gianseo_targets = []    # 노랑(Claude Max) 행 정보
    billing_schedule = []   # 서비스·결제일·금액 스케줄 (누락 계정 추정용)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        vals = [c.value for c in row]
        a = (str(vals[0]).strip() if vals[0] is not None else "")
        # 섹션 전환
        if a == "본부별 공용 유료 AI":
            section = 1; continue
        if a == "부설연구소" and (len(vals) < 2 or not vals[1]):
            section = 2; cur_name = ""; continue
        if a == "전직원" and (len(vals) < 2 or not vals[1]):
            section = 3; continue
        if a == "AX프로젝트":
            section = 4; cur_hq = ""; continue
        if a == "LAW":
            section = 5; cur_name = ""; continue
        if a in ("본부", "이름", "사용자", "사용먀"):     # 헤더행
            continue

        ai = vals[1] if len(vals) > 1 else None   # B 사용AI
        plan = vals[2] if len(vals) > 2 else None  # C 플랜
        idpw = vals[3] if len(vals) > 3 else None  # D 이메일/ID
        card = vals[7] if len(vals) > 7 else None  # H 카드
        plan_cell = row[2] if len(row) > 2 else None

        if section == 1:
            if a:
                cur_hq = a
            user = cur_hq
        elif section == 2:
            if a:
                cur_name = a
            user = f"부설연구소 {cur_name}".strip()
        elif section == 3:
            user = "회사공용"
        elif section == 4:
            # AX프로젝트: A열=본부(forward fill), B열=사용자 이름
            if a:
                cur_hq = a
            ax_name = str(vals[1]).strip() if vals[1] else ""
            user = f"AX {ax_name}" if ax_name else f"AX {cur_hq}"
        elif section == 5:
            # LAW: A열=사용자명(이름 그대로 사용)
            if a:
                cur_name = a
            user = cur_name
        else:
            continue

        email = str(idpw).strip() if idpw else ""
        if email and "@" in email:
            email_to_user[email.lower()] = user
        # 사람이름 → 사용자 (부설연구소 개인 등)
        if section == 2 and a:
            name_to_user[a.split()[0]] = user

        # 노랑 = Claude Max 기안서 대상
        if plan_cell is not None and _is_yellow(plan_cell):
            gianseo_targets.append({
                "user": user, "ai": ai, "plan": str(plan), "email": email,
                "card": str(card) if card else "",
            })

        # 결제 스케줄 추출 (누락 계정 추정용)
        billing_day_raw = vals[5] if len(vals) > 5 else None
        amount_raw = vals[4] if len(vals) > 4 else None
        bd_m = re.search(r"\d+", str(billing_day_raw)) if billing_day_raw is not None else None
        billing_day = int(bd_m.group()) if bd_m else None
        am_m = re.search(r"[\d,]+", str(amount_raw)) if amount_raw is not None else None
        amount_approx = float(am_m.group().replace(",", "")) if am_m else None

        if section == 4:
            svc_key = "ANTHROPIC/CLAUDE"  # AX프로젝트 = 전부 Claude Max
        else:
            ai_s = str(ai).lower() if ai else ""
            if "gemini" in ai_s:
                svc_key = "GOOGLE PLAY" if (amount_approx and amount_approx > 1000) else "GOOGLE"
            elif "claude" in ai_s:
                svc_key = "ANTHROPIC/CLAUDE"
            elif "gpt" in ai_s or "openai" in ai_s:
                svc_key = "OPENAI"
            elif "google api" in ai_s or ("google" in ai_s):
                svc_key = "GOOGLE CLOUD"
            elif "supabase" in ai_s:
                svc_key = "SUPABASE"
            elif "midjourney" in ai_s:
                svc_key = "MIDJOURNEY"
            else:
                svc_key = ""

        if svc_key and user:
            billing_schedule.append({
                "user": user,
                "service_key": svc_key,
                "amount_approx": amount_approx,
                "billing_day": billing_day,
            })

    merged_aliases = dict(DEFAULT_ALIASES)
    if aliases:
        merged_aliases.update(aliases)
    return {
        "email_to_user": email_to_user,
        "name_to_user": name_to_user,
        "aliases": merged_aliases,
        "gianseo_targets": gianseo_targets,
        "billing_schedule": billing_schedule,
    }


def resolve_user(master, email, payee=None):
    """이메일/ID/받는사람으로 사용자표기 찾기."""
    if (not email) and payee:
        nm = str(payee).strip().split()[0] if str(payee).strip() else ""
        for k, u in master.get("name_to_user", {}).items():
            if nm and (nm == k or nm in k or k in nm):
                return u
    if not email:
        return "회사 공용계정"
    e = email.lower().strip()
    if e in master["email_to_user"]:
        return master["email_to_user"][e]
    local = e.split("@")[0]
    # 별칭(부분일치 허용: towyssl@gmail/naver 모두)
    for key, user in master["aliases"].items():
        if local == key or local.startswith(key) or key in local:
            return user
    # local-part가 시트 이메일의 local과 같으면
    for se, u in master["email_to_user"].items():
        if se.split("@")[0] == local:
            return u
    if payee:
        nm = str(payee).strip().split()[0] if str(payee).strip() else ""
        for k, u in master.get("name_to_user", {}).items():
            if nm and (nm == k or nm in k or k in nm):
                return u
    return "회사 공용계정"


if __name__ == "__main__":
    m = load_master(sys.argv[1])
    print("=== email → 사용자 ===")
    for e, u in m["email_to_user"].items():
        print(f"  {e:28} → {u}")
    print("=== 별칭 ===")
    for k, v in m["aliases"].items():
        print(f"  {k:12} → {v}")
    print("=== 기안서 대상(노랑) ===")
    for t in m["gianseo_targets"]:
        print(f"  {t['user']:24} | {t['ai']} {t['plan']} | {t['email']} | card {t['card']}")
